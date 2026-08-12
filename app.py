
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from pathlib import Path
from werkzeug.utils import secure_filename
import openpyxl, sqlite3, io, zipfile, re, shutil, json, math
from collections import defaultdict

BASE = Path(__file__).resolve().parent
DATA = BASE / "data"
UPLOADS = DATA / "uploads"
OUTPUTS = BASE / "outputs"
DB = DATA / "master.db"
UPLOADS.mkdir(parents=True, exist_ok=True)
OUTPUTS.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.secret_key = "invoice-consolidation-web-v1"

def db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    con.execute("""CREATE TABLE IF NOT EXISTS master_items(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hs_code TEXT UNIQUE NOT NULL,
        product_code TEXT DEFAULT '',
        description TEXT DEFAULT ''
    )""")
    con.commit()
    return con

def normalize(v):
    if v is None: return ""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit(): s = s[:-2]
    return s

def load_master(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["MASTER LIST"] if "MASTER LIST" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows=[]
    for r in range(2, ws.max_row+1):
        hs=normalize(ws.cell(r,1).value)
        if hs:
            rows.append((hs, normalize(ws.cell(r,2).value), ""))
    con=db()
    con.executemany("""INSERT INTO master_items(hs_code,product_code,description)
                       VALUES(?,?,?)
                       ON CONFLICT(hs_code) DO UPDATE SET product_code=excluded.product_code""", rows)
    con.commit(); con.close()
    return len(rows)

def master_map():
    con=db()
    rows=con.execute("SELECT hs_code, product_code FROM master_items").fetchall()
    con.close()
    return {r["hs_code"]: r["product_code"] for r in rows}

def find_invoice_header(ws):
    for r in range(1, min(ws.max_row, 100)+1):
        vals=[normalize(ws.cell(r,c).value).lower() for c in range(1, ws.max_column+1)]
        if "hs code" in vals and ("sales invoice number" in vals or "invoice number" in vals):
            return r, {v:i+1 for i,v in enumerate(vals) if v}
    return None, {}

def read_invoice(path):
    wb=openpyxl.load_workbook(path, data_only=True)
    ws=wb.active
    hr, cols=find_invoice_header(ws)
    if not hr:
        raise ValueError("Could not find the invoice detail header. Required columns include HS Code and Sales Invoice Number.")
    def col(*names):
        for n in names:
            if n.lower() in cols: return cols[n.lower()]
        return None
    c_hs=col("hs code","cccn number")
    c_origin=col("country of origin","coo")
    c_desc=col("item description","description")
    c_amount=col("total amount including charges","total amount")
    c_qty=col("qty","quantity")
    c_net=col("nett weight","net weight")
    c_gross=col("gross weight")
    c_uom=col("unit")
    c_inv=col("sales invoice number","invoice number")
    c_currency=col("currency")
    c_part=col("part no","product code")
    if not all([c_hs,c_origin,c_desc,c_amount,c_inv]):
        raise ValueError("Missing required invoice columns.")
    rows=[]
    for r in range(hr+1, ws.max_row+1):
        hs=normalize(ws.cell(r,c_hs).value)
        inv=normalize(ws.cell(r,c_inv).value)
        if not hs or not inv: continue
        amount=ws.cell(r,c_amount).value
        try: amount=float(amount or 0)
        except: amount=0.0
        qty=ws.cell(r,c_qty).value if c_qty else ""
        try: qty=float(qty or 0)
        except: qty=0.0
        net=ws.cell(r,c_net).value if c_net else 0
        try: net=float(net or 0)
        except: net=0.0
        gross=ws.cell(r,c_gross).value if c_gross else 0
        try: gross=float(gross or 0)
        except: gross=0.0
        rows.append({
            "hs":hs, "origin":normalize(ws.cell(r,c_origin).value),
            "desc":normalize(ws.cell(r,c_desc).value),
            "amount":amount, "qty":qty, "net":net, "gross":gross,
            "uom":normalize(ws.cell(r,c_uom).value) if c_uom else "",
            "invoice":inv, "part":normalize(ws.cell(r,c_part).value) if c_part else "",
            "currency":normalize(ws.cell(r,c_currency).value) if c_currency else ""
        })
    return rows

def consolidate(rows):
    mm=master_map()
    groups={}
    for x in rows:
        key=(x["hs"],x["origin"])
        if key not in groups:
            groups[key]={**x, "descs":[], "invoices":set(), "amount":0.0, "qty":0.0, "net":0.0, "gross":0.0}
        g=groups[key]
        if x["desc"] and x["desc"] not in g["descs"]: g["descs"].append(x["desc"])
        g["invoices"].add(x["invoice"])
        g["amount"] += x["amount"]
        g["qty"] += x["qty"]
        g["net"] += x["net"]
        g["gross"] += x["gross"]
    out=[]
    for g in groups.values():
        hs=g["hs"]
        pc=mm.get(hs,"")
        typ=pc[:1].upper() if pc else ""
        out.append({
            "hs":hs, "origin":g["origin"], "desc":" / ".join(g["descs"]),
            "amount":round(g["amount"],6), "qty":g["qty"], "net":g["net"], "gross":g["gross"],
            "invoice":sorted(g["invoices"], key=str), "product_code":pc, "type":typ
        })
    return sorted(out,key=lambda x:(x["type"],x["origin"],x["hs"]))

def fill_template(template_path, items, header):
    wb=openpyxl.load_workbook(template_path)
    if "IMP_BL" not in wb.sheetnames or "Invoice" not in wb.sheetnames:
        raise ValueError("Template must contain IMP_BL and Invoice sheets.")
    imp=wb["IMP_BL"]; invws=wb["Invoice"]
    fields=["Inward MAWB","Inward HAWB","Arrival Date","Inward Flight / Voyage Number",
            "Inward Vessel Name","Loading Port","Load Port Code","Total Gross Wt",
            "Gross Weight UOM","Outer Pack Count","Out Pack UOM"]
    for i,f in enumerate(fields,1):
        imp.cell(i,2).value=header.get(f,"")
    invs=sorted({n for x in items for n in x["invoice"]}, key=str)
    imp["B12"]="Inv no : " + ", ".join(invs)
    # clear old data after row 1
    if invws.max_row > 1:
        invws.delete_rows(2, invws.max_row-1)
    for x in items:
        row=[
            ", ".join(x["invoice"]), "", x.get("currency",""), x["amount"], "",
            x["desc"], "", x["hs"], x["qty"], x["amount"], x["origin"], x["net"],
            "", "", x["product_code"], "", "", ""
        ]
        invws.append(row)
    return wb

@app.route("/")
def index():
    con=db()
    count=con.execute("SELECT COUNT(*) c FROM master_items").fetchone()["c"]
    con.close()
    return render_template("index.html", master_count=count)

@app.post("/process")
def process():
    invoice=request.files.get("invoice")
    master=request.files.get("master")
    template=request.files.get("template")
    if not invoice or not master or not template:
        flash("Please select Invoice, Master List and Output Template.")
        return redirect(url_for("index"))
    inv_path=UPLOADS/secure_filename(invoice.filename)
    mst_path=UPLOADS/secure_filename(master.filename)
    tpl_path=UPLOADS/secure_filename(template.filename)
    invoice.save(inv_path); master.save(mst_path); template.save(tpl_path)
    try:
        load_master(mst_path)
        rows=read_invoice(inv_path)
        items=consolidate(rows)
        session_file=DATA/"last_items.json"
        session_file.write_text(json.dumps(items, ensure_ascii=False))
        types=sorted({x["type"] for x in items if x["type"]})
        noncontrol=sum(1 for x in items if not x["product_code"])
        unmatched=sum(1 for x in items if x["hs"] not in master_map())
        return render_template("group.html", types=types, noncontrol=noncontrol, unmatched=unmatched,
                               count=len(items), invoice_count=len({n for x in items for n in x["invoice"]}),
                               template=str(tpl_path.name))
    except Exception as e:
        flash(str(e)); return redirect(url_for("index"))

@app.post("/generate")
def generate():
    items=json.loads((DATA/"last_items.json").read_text())
    template_name=request.form.get("template")
    template=UPLOADS/template_name
    header={k:request.form.get(k,"") for k in [
        "Inward MAWB","Inward HAWB","Arrival Date","Inward Flight / Voyage Number",
        "Inward Vessel Name","Loading Port","Load Port Code","Total Gross Wt",
        "Gross Weight UOM","Outer Pack Count","Out Pack UOM"]}
    selected=request.form.getlist("types")
    if "NONCONTROL" in selected:
        selected_items=[x for x in items if not x["product_code"]]
    else:
        selected_items=[x for x in items if x["type"] in selected]
    split=request.form.get("split")=="origin"
    buckets=defaultdict(list)
    if split:
        for x in selected_items: buckets[x["origin"]].append(x)
    else:
        buckets["ALL"]=selected_items
    files=[]
    for bucket, arr in buckets.items():
        for start in range(0,len(arr),50):
            chunk=arr[start:start+50]
            wb=fill_template(template,chunk,header)
            safe=re.sub(r"[^A-Za-z0-9_-]+","_",bucket or "ALL")
            idx=start//50+1
            name=f"Consolidated_{safe}_{idx:03d}.xlsx"
            p=OUTPUTS/name
            wb.save(p); files.append(p)
    if not files:
        flash("No items matched the selected product groups.")
        return redirect(url_for("index"))
    if len(files)==1:
        return send_file(files[0],as_attachment=True,download_name=files[0].name)
    z=OUTPUTS/"Invoice_Consolidation_Outputs.zip"
    with zipfile.ZipFile(z,"w",zipfile.ZIP_DEFLATED) as zz:
        for p in files: zz.write(p,p.name)
    return send_file(z,as_attachment=True,download_name=z.name)

@app.route("/master")
def master():
    q=request.args.get("q","")
    con=db()
    if q:
        rows=con.execute("SELECT * FROM master_items WHERE hs_code LIKE ? OR product_code LIKE ? ORDER BY hs_code",
                         (f"%{q}%",f"%{q}%")).fetchall()
    else:
        rows=con.execute("SELECT * FROM master_items ORDER BY hs_code").fetchall()
    con.close()
    return render_template("master.html", rows=rows, q=q)

@app.post("/master/save")
def master_save():
    hs=normalize(request.form.get("hs_code"))
    pc=normalize(request.form.get("product_code"))
    desc=normalize(request.form.get("description"))
    if not hs:
        flash("HS Code is required.")
    else:
        con=db()
        con.execute("""INSERT INTO master_items(hs_code,product_code,description) VALUES(?,?,?)
                       ON CONFLICT(hs_code) DO UPDATE SET product_code=excluded.product_code,
                       description=excluded.description""",(hs,pc,desc))
        con.commit(); con.close()
        flash("Master List saved.")
    return redirect(url_for("master"))

@app.post("/master/delete/<int:item_id>")
def master_delete(item_id):
    con=db(); con.execute("DELETE FROM master_items WHERE id=?",(item_id,)); con.commit(); con.close()
    flash("Master item deleted.")
    return redirect(url_for("master"))

@app.post("/master/import")
def master_import():
    f=request.files.get("master")
    if not f: flash("Select a Master List Excel file.")
    else:
        p=UPLOADS/secure_filename(f.filename); f.save(p)
        try: flash(f"Imported {load_master(p)} Master List rows.")
        except Exception as e: flash(str(e))
    return redirect(url_for("master"))

if __name__=="__main__":
    db().close()
    app.run(host="0.0.0.0", port=5000, debug=False)
